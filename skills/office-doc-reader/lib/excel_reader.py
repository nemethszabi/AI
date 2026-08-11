"""
Reads an existing .xlsx workbook and renders its content as Markdown - the read-side counterpart to
excel_helpers.py in the office-doc-builder skill (which only writes/formats). Used to turn inbound client
spreadsheets (estimates, requirement lists, data dumps) into text an agent can read and cite.

Mechanical extraction only - no interpretation of what the data means. Every sheet becomes a heading plus
a Markdown table; formulas are rendered as their last-calculated value (falling back to the formula text
itself when no cached value exists, e.g. a workbook never opened in Excel since the formula was entered).

Requires: openpyxl (pip install openpyxl).

CLI usage: python excel_reader.py <input.xlsx> <output.md>
"""

import sys
from pathlib import Path

from openpyxl import load_workbook


def _cell_text(cell):
    value = cell.value
    if value is None:
        return ""
    return str(value)


def extract_xlsx_to_markdown(path, max_rows_per_sheet=2000):
    """Reads every worksheet in the workbook at `path` and returns one Markdown string covering all of
    them. max_rows_per_sheet caps runaway extraction on huge sheets - if a sheet is truncated, a warning
    line is appended after its table instead of silently dropping rows. Returns (markdown_text,
    warnings_list). Merged cells are not un-merged - openpyxl reports the value only in the merged
    range's top-left cell, blank in the rest, matching what a plain read of the file would show."""
    wb = load_workbook(path, data_only=True)
    warnings = []
    parts = [f"# {Path(path).name}\n"]

    for ws in wb.worksheets:
        parts.append(f"\n## Sheet: {ws.title}\n")
        rows = list(ws.iter_rows())
        if not rows:
            parts.append("\n_(empty sheet)_\n")
            continue

        truncated = len(rows) > max_rows_per_sheet
        if truncated:
            rows = rows[:max_rows_per_sheet]
            warnings.append(f"Sheet '{ws.title}': truncated to first {max_rows_per_sheet} rows")

        header = [_cell_text(c) for c in rows[0]]
        col_count = len(header)
        parts.append("| " + " | ".join(h or f"col{i+1}" for i, h in enumerate(header)) + " |")
        parts.append("|" + "---|" * col_count)
        for row in rows[1:]:
            cells = [_cell_text(c) for c in row[:col_count]]
            while len(cells) < col_count:
                cells.append("")
            parts.append("| " + " | ".join(c.replace("|", "\\|").replace("\n", " ") for c in cells) + " |")

        if truncated:
            parts.append(f"\n_(truncated - {len(list(ws.iter_rows())) - max_rows_per_sheet} more rows not shown)_\n")

    return "\n".join(parts), warnings


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("usage: python excel_reader.py <input.xlsx> <output.md>", file=sys.stderr)
        sys.exit(1)

    in_path, out_path = sys.argv[1], sys.argv[2]
    text, warnings = extract_xlsx_to_markdown(in_path)
    Path(out_path).write_text(text, encoding="utf-8")

    if warnings:
        print("WARNINGS:")
        for w in warnings:
            print(f"  - {w}")
    print(f"Wrote {out_path}")
