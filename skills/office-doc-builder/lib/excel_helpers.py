"""
Reusable openpyxl formatting helpers - extracted from patterns proven in an earlier Excel-generation
script and generalized, so future Excel-generating skills/scripts don't re-derive the same styling logic
from scratch each time.

Import these into a short one-off script rather than writing raw openpyxl styling inline - keeps visual
consistency across every document this framework generates, and keeps the actual per-task script short
(just the data + which helpers to call).

Requires: openpyxl (pip install openpyxl).
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Reusable named colors (ARGB hex) - not exhaustive, add more here as new needs come up rather than
# inlining a fresh color per script.
COLOR_HEADER_BLUE = "FF4472C4"
COLOR_TOTAL_YELLOW = "FFFFFF00"
COLOR_HIGHLIGHT_GREEN = "FF92D050"
COLOR_BORDER_GRAY = "FFB0B0B0"

THIN_BORDER = Border(
    left=Side(style="thin", color=COLOR_BORDER_GRAY),
    right=Side(style="thin", color=COLOR_BORDER_GRAY),
    top=Side(style="thin", color=COLOR_BORDER_GRAY),
    bottom=Side(style="thin", color=COLOR_BORDER_GRAY),
)


def styled_header_row(ws, row, headers, col_widths=None, fill_color=COLOR_HEADER_BLUE, start_col=1):
    """Writes a header row with white-on-color fill, bold font, centered+wrapped text, and sets column
    widths if given. headers: list of strings. col_widths: optional list of ints, same length as headers.
    Returns the row number written (for chaining)."""
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(bold=True, color="FFFFFFFF")
    for i, name in enumerate(headers, start=start_col):
        cell = ws.cell(row=row, column=i, value=name)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        if col_widths:
            ws.column_dimensions[get_column_letter(i)].width = col_widths[i - start_col]
    return row


def write_table(ws, start_row, rows, start_col=1, number_format=None, borders=True):
    """Writes a 2D list of row-values starting at (start_row, start_col), one row per list entry.
    number_format: optional dict {col_offset: format_string} to apply a number format to specific
    columns (0-indexed relative to start_col), e.g. {1: "#,##0"} formats the second column as an integer
    with thousands separators. Returns the row number after the last row written."""
    row = start_row
    for row_values in rows:
        for i, value in enumerate(row_values):
            cell = ws.cell(row=row, column=start_col + i, value=value)
            if borders:
                cell.border = THIN_BORDER
            if number_format and i in number_format:
                cell.number_format = number_format[i]
        row += 1
    return row


def total_row(ws, row, label, sum_ranges, label_col=1, fill_color=COLOR_TOTAL_YELLOW):
    """Writes a bold total row. sum_ranges: dict {col: (first_row, last_row)} - for each entry, writes
    a =SUM(colFirst:colLast) formula in that column at this row. Returns the row written (for chaining)."""
    ws.cell(row=row, column=label_col, value=label).font = Font(bold=True)
    ws.cell(row=row, column=label_col).fill = PatternFill(
        start_color=fill_color, end_color=fill_color, fill_type="solid"
    )
    for col, (first_row, last_row) in sum_ranges.items():
        col_letter = get_column_letter(col)
        cell = ws.cell(row=row, column=col, value=f"=SUM({col_letter}{first_row}:{col_letter}{last_row})")
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.number_format = "#,##0"
    return row


def highlight_cell(cell, fill_color=COLOR_HIGHLIGHT_GREEN):
    """Applies a highlight fill to a single cell - e.g. marking a value that needs manual fill-in later."""
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")


def freeze_header(ws, header_row):
    """Freezes panes so everything at or above header_row stays visible while scrolling."""
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate


def set_column_widths(ws, widths, start_col=1):
    """widths: list of ints. Sets column widths starting at start_col."""
    for i, w in enumerate(widths, start=start_col):
        ws.column_dimensions[get_column_letter(i)].width = w
